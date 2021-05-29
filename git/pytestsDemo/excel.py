import openpyxl
book = openpyxl.load_workbook("C:\\Users\\mhatr\\PycharmProjects\\SeleniumProject\\exceldata.xlsx")
sheet = book.active

dict = {}

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column):
        #print(sheet.cell(row= i, column= j).value)
        if (sheet.cell(row=i, column=1).value == "01"):
            dict[sheet.cell(row=1, column=j+1).value] = sheet.cell(row=i, column=j+1).value

print([dict])